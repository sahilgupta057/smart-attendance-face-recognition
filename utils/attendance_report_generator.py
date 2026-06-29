import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os
from pathlib import Path
import tempfile

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    HRFlowable,
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4, landscape


# ── Shared PDF colours (match Excel palette) ─────────────────────────────────
_HEADER_BG   = colors.HexColor("#1F4E78")   # dark blue header
_PRESENT_BG  = colors.HexColor("#C6EFCE")   # light green
_PRESENT_FG  = colors.HexColor("#006100")   # dark green text
_ABSENT_BG   = colors.HexColor("#FFC7CE")   # light red
_ABSENT_FG   = colors.HexColor("#9C0006")   # dark red text
_SUMMARY_BG  = colors.HexColor("#E7E6E6")   # light grey summary
_PCT_OK_BG   = colors.HexColor("#C6EFCE")   # ≥ 75 %
_PCT_LOW_BG  = colors.HexColor("#FFC7CE")   # < 75 %
_WHITE       = colors.white
_BLACK       = colors.black


class AttendanceReportGenerator:
    """Generate Excel and PDF reports from attendance data."""

    def __init__(self, username=None, account_id=None):
        self.username   = username
        self.account_id = account_id
        base = f"{username}" if username else ""
        if account_id is not None:
            base = f"{base}_{account_id}"
        self.reports_dir = os.path.join("reports", base) if base else "reports"
        os.makedirs(self.reports_dir, exist_ok=True)

    # =========================================================================
    #  EXCEL — DAILY
    # =========================================================================
    def generate_daily_report(self, attendance_data, date_str=None,
                              department=None, subject=None):
        if date_str is None:
            date_str = datetime.now().strftime("%Y-%m-%d")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

        header_fill  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font  = Font(bold=True, color="FFFFFF", size=12)
        present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        absent_fill  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value     = f"Attendance Report - {date_str}"
        title_cell.font      = Font(bold=True, size=14, color="1F4E78")
        title_cell.alignment = center_align

        ws.merge_cells('A2:G2')
        metadata_cell = ws['A2']
        metadata_cell.value     = f"Department: {department or 'All'} | Subject: {subject or 'All'}"
        metadata_cell.font      = Font(italic=True, size=10)
        metadata_cell.alignment = center_align

        headers = ["S.No", "Name", "Roll", "Department", "Year", "Time", "Status"]
        for col_num, header in enumerate(headers, 1):
            cell           = ws.cell(row=4, column=col_num)
            cell.value     = header
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = border

        for row_num, record in enumerate(attendance_data, 5):
            ws.cell(row=row_num, column=1).value = row_num - 4
            ws.cell(row=row_num, column=2).value = record.get("name",   "")
            ws.cell(row=row_num, column=3).value = record.get("roll",   "")
            ws.cell(row=row_num, column=4).value = record.get("dept",   "")
            ws.cell(row=row_num, column=5).value = record.get("year",   "")
            ws.cell(row=row_num, column=6).value = record.get("time",   "")
            ws.cell(row=row_num, column=7).value = record.get("status", "")

            status = record.get("status", "").lower()
            for col in range(1, 8):
                cell           = ws.cell(row=row_num, column=col)
                cell.border    = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col == 7:
                    if status == "present":
                        cell.fill = present_fill
                        cell.font = Font(bold=True, color="006100")
                    elif status == "absent":
                        cell.fill = absent_fill
                        cell.font = Font(bold=True, color="9C0006")

        summary_row  = len(attendance_data) + 6
        present_count = sum(1 for r in attendance_data if r.get("status","").lower() == "present")
        absent_count  = len(attendance_data) - present_count

        ws.merge_cells(f'A{summary_row}:G{summary_row}')
        summary_cell           = ws[f'A{summary_row}']
        summary_cell.value     = (f"Total Students: {len(attendance_data)} | "
                                   f"Present: {present_count} | Absent: {absent_count}")
        summary_cell.font      = Font(bold=True, size=11)
        summary_cell.fill      = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        summary_cell.alignment = center_align

        safe_date      = date_str.replace("-", "")
        dept_suffix    = f"_{department.replace(' ','_')}" if department else ""
        subject_suffix = f"_{subject.replace(' ','_')}"    if subject    else ""
        filename       = f"Attendance_{safe_date}{dept_suffix}{subject_suffix}.xlsx"
        filepath       = os.path.join(self.reports_dir, filename)
        wb.save(filepath)
        return filepath

    # =========================================================================
    #  PDF — DAILY  (mirrors Excel daily layout)
    # =========================================================================
    def generate_daily_pdf_report(self, attendance_data, date_str,
                                  department=None, subject=None):
        """
        Daily PDF report — columns and styling identical to the Excel version:
        S.No | Name | Roll | Department | Year | Time | Status
        Includes metadata row and colour-coded summary footer.
        """
        safe_date = date_str.replace("-", "")
        dept_suffix = f"_{department.replace(' ','_')}" if department else ""
        filename  = f"Attendance_{safe_date}{dept_suffix}.pdf"
        filepath  = os.path.join(self.reports_dir, filename)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            leftMargin=0.5*inch, rightMargin=0.5*inch,
            topMargin=0.6*inch,  bottomMargin=0.6*inch,
        )

        styles   = getSampleStyleSheet()
        elements = []

        # ── Title ─────────────────────────────────────────────────────────────
        title_style = ParagraphStyle(
            "rpt_title",
            parent=styles["Title"],
            textColor=colors.HexColor("#1F4E78"),
            fontSize=14,
            spaceAfter=4,
        )
        elements.append(Paragraph(f"Attendance Report - {date_str}", title_style))

        # ── Metadata ──────────────────────────────────────────────────────────
        meta_style = ParagraphStyle(
            "rpt_meta",
            parent=styles["Normal"],
            fontSize=9,
            textColor=colors.HexColor("#555555"),
            spaceAfter=10,
        )
        elements.append(Paragraph(
            f"Department: {department or 'All'} &nbsp;|&nbsp; Subject: {subject or 'All'}",
            meta_style
        ))
        elements.append(HRFlowable(width="100%", thickness=0.5,
                                   color=colors.HexColor("#1F4E78")))
        elements.append(Spacer(1, 8))

        # ── Table data ────────────────────────────────────────────────────────
        col_headers = ["S.No", "Name", "Roll", "Department", "Year", "Time", "Status"]
        table_data  = [col_headers]

        for idx, record in enumerate(attendance_data, 1):
            table_data.append([
                str(idx),
                record.get("name",   ""),
                record.get("roll",   ""),
                record.get("dept",   ""),
                str(record.get("year",   "")),
                record.get("time",   ""),
                record.get("status", ""),
            ])

        # Summary footer row
        present_count = sum(1 for r in attendance_data
                            if r.get("status", "").lower() == "present")
        absent_count  = len(attendance_data) - present_count
        summary_text  = (f"Total: {len(attendance_data)}   "
                         f"Present: {present_count}   "
                         f"Absent: {absent_count}")
        table_data.append(["", summary_text, "", "", "", "", ""])

        # ── Column widths (points, page ≈ 7.5 in usable) ─────────────────────
        col_widths = [0.45*inch, 1.8*inch, 0.9*inch,
                      1.3*inch, 0.55*inch, 0.8*inch, 0.8*inch]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        last_data_row = len(attendance_data)   # 0-based: rows 1..last_data_row are data
        summary_row_idx = last_data_row + 1    # last row in table_data

        style_cmds = [
            # Header row
            ("BACKGROUND",  (0, 0), (-1, 0),  _HEADER_BG),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  _WHITE),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0),  9),
            ("ALIGN",       (0, 0), (-1, 0),  "CENTER"),
            ("VALIGN",      (0, 0), (-1, 0),  "MIDDLE"),

            # Data rows
            ("FONTNAME",    (0, 1), (-1, last_data_row), "Helvetica"),
            ("FONTSIZE",    (0, 1), (-1, last_data_row), 8),
            ("ALIGN",       (0, 1), (-1, last_data_row), "CENTER"),
            ("VALIGN",      (0, 1), (-1, last_data_row), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, last_data_row),
             [colors.white, colors.HexColor("#EBF3FB")]),

            # Summary footer
            ("SPAN",        (0, summary_row_idx), (-1, summary_row_idx)),
            ("BACKGROUND",  (0, summary_row_idx), (-1, summary_row_idx), _SUMMARY_BG),
            ("FONTNAME",    (0, summary_row_idx), (-1, summary_row_idx), "Helvetica-Bold"),
            ("FONTSIZE",    (0, summary_row_idx), (-1, summary_row_idx), 9),
            ("ALIGN",       (0, summary_row_idx), (-1, summary_row_idx), "CENTER"),

            # Grid
            ("GRID",        (0, 0), (-1, last_data_row), 0.4, _BLACK),
            ("LINEBELOW",   (0, summary_row_idx), (-1, summary_row_idx), 0.4, _BLACK),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ]

        # Per-row status colour for Status column (index 6)
        for row_idx, record in enumerate(attendance_data, 1):
            status = record.get("status", "").lower()
            if status == "present":
                style_cmds += [
                    ("BACKGROUND", (6, row_idx), (6, row_idx), _PRESENT_BG),
                    ("TEXTCOLOR",  (6, row_idx), (6, row_idx), _PRESENT_FG),
                    ("FONTNAME",   (6, row_idx), (6, row_idx), "Helvetica-Bold"),
                ]
            elif status == "absent":
                style_cmds += [
                    ("BACKGROUND", (6, row_idx), (6, row_idx), _ABSENT_BG),
                    ("TEXTCOLOR",  (6, row_idx), (6, row_idx), _ABSENT_FG),
                    ("FONTNAME",   (6, row_idx), (6, row_idx), "Helvetica-Bold"),
                ]

        table.setStyle(TableStyle(style_cmds))
        elements.append(table)

        doc.build(elements)
        return filepath

    # =========================================================================
    #  EXCEL — MONTHLY
    # =========================================================================
    def generate_monthly_report(self, attendance_data, year=None, month=None,
                                department=None, student_name=None,
                                save_to_reports_dir=True):
        if year  is None: year  = datetime.now().year
        if month is None: month = datetime.now().month

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly"

        ws.column_dimensions['A'].width = 25

        header_fill  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font  = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        unique_dates    = sorted(set(r.get("date", "") for r in attendance_data))
        unique_students = {}
        for record in attendance_data:
            roll = record.get("roll", "")
            if roll and roll not in unique_students:
                unique_students[roll] = {
                    "name": record.get("name", ""),
                    "dept": record.get("dept", ""),
                    "year": record.get("year", "")
                }

        if student_name:
            unique_students = {
                roll: info for roll, info in unique_students.items()
                if info["name"] == student_name
            }

        month_name = datetime(year, month, 1).strftime("%B %Y")
        title_text = f"Monthly Attendance Report - {month_name}"
        if student_name:
            title_text += f" ({student_name})"

        ws.merge_cells('A1:L1')
        title_cell           = ws['A1']
        title_cell.value     = title_text
        title_cell.font      = Font(bold=True, size=14, color="1F4E78")
        title_cell.alignment = center_align

        ws.cell(row=3, column=1).value     = "Student Name"
        ws.cell(row=3, column=1).fill      = header_fill
        ws.cell(row=3, column=1).font      = header_font
        ws.cell(row=3, column=1).border    = border

        for col_num, date_str in enumerate(unique_dates, 2):
            cell           = ws.cell(row=3, column=col_num)
            cell.value     = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d %b %Y")
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15

        present_col    = len(unique_dates) + 2
        absent_col     = len(unique_dates) + 3
        percentage_col = len(unique_dates) + 4

        for col, label in [(present_col, "Present"),
                           (absent_col,  "Absent"),
                           (percentage_col, "Attendance %")]:
            c           = ws.cell(row=3, column=col)
            c.value     = label
            c.fill      = header_fill
            c.font      = header_font
            c.border    = border
            c.alignment = center_align

        for row_num, (roll, student_info) in enumerate(unique_students.items(), 4):
            ws.cell(row=row_num, column=1).value  = student_info["name"]
            ws.cell(row=row_num, column=1).border = border

            present_count = 0
            for date_col, date_str in enumerate(unique_dates, 2):
                status = "A"
                for record in attendance_data:
                    if record.get("roll") == roll and record.get("date") == date_str:
                        status = "P" if record.get("status","").lower() == "present" else "A"
                        if status == "P":
                            present_count += 1
                        break

                cell           = ws.cell(row=row_num, column=date_col)
                cell.value     = status
                cell.alignment = center_align
                cell.border    = border
                cell.fill      = (PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                  if status == "P" else
                                  PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))

            total_days  = len(unique_dates)
            absent_count = total_days - present_count
            pct          = round((present_count / total_days) * 100, 2) if total_days else 0

            for col, val in [(present_col, present_count),
                             (absent_col,  absent_count),
                             (percentage_col, f"{pct}%")]:
                c           = ws.cell(row=row_num, column=col)
                c.value     = val
                c.alignment = center_align
                c.border    = border
                c.font      = Font(bold=True)

            pct_cell      = ws.cell(row=row_num, column=percentage_col)
            pct_cell.fill = (PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                             if pct >= 75 else
                             PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))

        if student_name:
            safe_name = student_name.replace(" ", "_").replace("/", "_")
            filename  = f"Attendance_Monthly_{year}_{month:02d}_{safe_name}.xlsx"
        else:
            filename  = f"Attendance_Monthly_{year}_{month:02d}.xlsx"

        if save_to_reports_dir:
            filepath = os.path.join(self.reports_dir, filename)
            wb.save(filepath)
            return filepath
        else:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp_path = tmp.name
            tmp.close()
            wb.save(tmp_path)
            return tmp_path

    # =========================================================================
    #  PDF — MONTHLY  (mirrors Excel monthly layout)
    # =========================================================================
    def generate_monthly_pdf_report(self, attendance_data, year, month,
                                    department=None, student_name=None):
        """
        Monthly PDF report — columns and styling identical to the Excel version:
        Student Name | Date1 | Date2 | … | Present | Absent | Attendance %
        P/A cells are green/red; percentage cell is green ≥75 % else red.
        """
        month_name = datetime(year, month, 1).strftime("%B %Y")

        if student_name:
            safe_name = student_name.replace(" ", "_").replace("/", "_")
            filename  = f"Attendance_Monthly_{year}_{month:02d}_{safe_name}.pdf"
        else:
            filename  = f"Attendance_Monthly_{year}_{month:02d}.pdf"
        filepath = os.path.join(self.reports_dir, filename)

        # Many date columns → landscape A4
        doc = SimpleDocTemplate(
            filepath,
            pagesize=landscape(A4),
            leftMargin=0.4*inch, rightMargin=0.4*inch,
            topMargin=0.5*inch,  bottomMargin=0.5*inch,
        )

        styles   = getSampleStyleSheet()
        elements = []

        # ── Title ─────────────────────────────────────────────────────────────
        title_style = ParagraphStyle(
            "rpt_title_m",
            parent=styles["Title"],
            textColor=colors.HexColor("#1F4E78"),
            fontSize=14,
            spaceAfter=4,
        )
        title_text = f"Monthly Attendance Report - {month_name}"
        if student_name:
            title_text += f"  ({student_name})"
        elements.append(Paragraph(title_text, title_style))

        if department:
            meta_style = ParagraphStyle("rpt_meta_m", parent=styles["Normal"],
                                        fontSize=9, textColor=colors.HexColor("#555555"),
                                        spaceAfter=8)
            elements.append(Paragraph(f"Department: {department}", meta_style))

        elements.append(HRFlowable(width="100%", thickness=0.5,
                                   color=colors.HexColor("#1F4E78")))
        elements.append(Spacer(1, 8))

        # ── Build student & date index ────────────────────────────────────────
        unique_dates    = sorted(set(r.get("date", "") for r in attendance_data))
        unique_students = {}
        for record in attendance_data:
            roll = record.get("roll", "")
            if roll and roll not in unique_students:
                unique_students[roll] = {
                    "name": record.get("name", ""),
                    "dept": record.get("dept", ""),
                    "year": record.get("year", ""),
                }

        if student_name:
            unique_students = {
                roll: info for roll, info in unique_students.items()
                if info["name"] == student_name
            }

        # ── Table header row ──────────────────────────────────────────────────
        date_labels = [
            datetime.strptime(d, "%Y-%m-%d").strftime("%d %b")
            for d in unique_dates
        ]
        header_row  = ["Student Name"] + date_labels + ["Present", "Absent", "Att. %"]
        table_data  = [header_row]

        # Track cell colours: list of (row, col, bg_color)
        cell_colors = []
        header_col_count = len(header_row)

        for student_row_idx, (roll, info) in enumerate(unique_students.items(), 1):
            present_count = 0
            row_cells     = [info["name"]]

            for date_col_idx, date_str in enumerate(unique_dates):
                status = "A"
                for record in attendance_data:
                    if record.get("roll") == roll and record.get("date") == date_str:
                        status = "P" if record.get("status", "").lower() == "present" else "A"
                        if status == "P":
                            present_count += 1
                        break
                row_cells.append(status)
                # column index in table = 1 + date_col_idx (0-based date column)
                col_idx = 1 + date_col_idx
                bg = _PRESENT_BG if status == "P" else _ABSENT_BG
                cell_colors.append((student_row_idx, col_idx, bg))

            total_days   = len(unique_dates)
            absent_count = total_days - present_count
            pct          = round((present_count / total_days) * 100, 2) if total_days else 0

            row_cells += [str(present_count), str(absent_count), f"{pct}%"]

            # Summary columns: last three
            pct_col_idx = header_col_count - 1
            pct_bg      = _PCT_OK_BG if pct >= 75 else _PCT_LOW_BG
            cell_colors.append((student_row_idx, pct_col_idx, pct_bg))

            table_data.append(row_cells)

        # ── Column widths ─────────────────────────────────────────────────────
        # Landscape A4 usable ≈ 11.0 in
        name_w    = 1.6 * inch
        summary_w = 0.55 * inch
        n_dates   = len(unique_dates)
        # Remaining width shared among date columns
        usable    = 10.8 * inch
        date_w    = max(
            0.45 * inch,
            (usable - name_w - 3 * summary_w) / max(n_dates, 1)
        )
        col_widths = [name_w] + [date_w] * n_dates + [summary_w, summary_w, summary_w]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        last_data = len(unique_students)  # last data row index (1-based)

        base_cmds = [
            # Header
            ("BACKGROUND",   (0, 0), (-1, 0),          _HEADER_BG),
            ("TEXTCOLOR",    (0, 0), (-1, 0),          _WHITE),
            ("FONTNAME",     (0, 0), (-1, 0),          "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, 0),          8),
            ("ALIGN",        (0, 0), (-1, 0),          "CENTER"),
            ("VALIGN",       (0, 0), (-1, 0),          "MIDDLE"),

            # Data rows
            ("FONTNAME",     (0, 1), (-1, last_data),  "Helvetica"),
            ("FONTSIZE",     (0, 1), (-1, last_data),  8),
            ("ALIGN",        (0, 1), (-1, last_data),  "CENTER"),
            ("VALIGN",       (0, 1), (-1, last_data),  "MIDDLE"),
            # Name column left-aligned
            ("ALIGN",        (0, 1), (0, last_data),   "LEFT"),
            # Alternating row background (will be overridden by cell_colors per cell)
            ("ROWBACKGROUNDS", (0, 1), (-1, last_data),
             [colors.white, colors.HexColor("#EBF3FB")]),

            # Summary columns: bold
            ("FONTNAME",     (-3, 1), (-1, last_data), "Helvetica-Bold"),

            # Grid
            ("GRID",         (0, 0), (-1, last_data),  0.4, _BLACK),
            ("TOPPADDING",   (0, 0), (-1, -1),         3),
            ("BOTTOMPADDING",(0, 0), (-1, -1),         3),
            ("LEFTPADDING",  (0, 0), (-1, -1),         3),
            ("RIGHTPADDING", (0, 0), (-1, -1),         3),
        ]

        # Apply per-cell colours
        for (r, c, bg) in cell_colors:
            base_cmds.append(("BACKGROUND", (c, r), (c, r), bg))
            # Make P green text, A red text
            if c < header_col_count - 3:   # date columns only
                fg = _PRESENT_FG if bg == _PRESENT_BG else _ABSENT_FG
                base_cmds.append(("TEXTCOLOR", (c, r), (c, r), fg))
                base_cmds.append(("FONTNAME",  (c, r), (c, r), "Helvetica-Bold"))

        table.setStyle(TableStyle(base_cmds))
        elements.append(table)

        doc.build(elements)
        return filepath

    # =========================================================================
    #  UTILITIES
    # =========================================================================
    def get_all_reports(self):
        if not os.path.exists(self.reports_dir):
            return []
        reports = []
        for file in os.listdir(self.reports_dir):
            filepath     = os.path.join(self.reports_dir, file)
            reports.append({
                'filename': file,
                'filepath': filepath,
                'size':     os.path.getsize(filepath),
                'modified': datetime.fromtimestamp(os.path.getmtime(filepath)),
            })
        return sorted(reports, key=lambda x: x['modified'], reverse=True)

    def delete_report(self, filepath):
        if os.path.exists(filepath):
            os.remove(filepath)
            return True
        return False